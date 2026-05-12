import pickle, numpy as np, pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import os
os.chdir(os.path.dirname(os.path.abspath(__file__)))

with open('stats.pkl','rb') as f:
    data = pickle.load(f)

df=data['df']; N=data['N']; overall_rate=data['overall_rate']
age_stats=data['age_stats']; diag_stats=data['diag_stats']
los_stats=data['los_stats']; med_stats=data['med_stats']
a1c_stats=data['a1c_stats']; race_stats=data['race_stats']
inpat_stats=data['inpat_stats']

wb = Workbook()

# ── Style helpers ─────────────────────────────────────────────────────────────
NAVY='1E3A5F'; BLUE='2563EB'; RED='DC2626'; AMBER='D97706'
GREEN='059669'; LIGHT='EFF6FF'; WHITE='FFFFFF'
LGRAY='F1F5F9'; MGRAY='CBD5E1'; DKGRAY='475569'; SLATE='64748B'

def fill(h): return PatternFill('solid', fgColor=h)
def hf(sz=11,bold=True,color=WHITE): return Font(name='Arial',bold=bold,size=sz,color=color)
def bf(sz=10,bold=False,color='1E293B'): return Font(name='Arial',bold=bold,size=sz,color=color)
def bdr(style='thin',color='CBD5E1'):
    s=Side(style=style,color=color); return Border(left=s,right=s,top=s,bottom=s)
def ctr(wrap=False): return Alignment(horizontal='center',vertical='center',wrap_text=wrap)
def pct(v): return f'{v:.1f}%' if not (isinstance(v,float) and np.isnan(v)) else 'N/A'

def section_hdr(ws,row,c1,c2,text,color=BLUE):
    ws.merge_cells(f'{c1}{row}:{c2}{row}')
    c=ws[f'{c1}{row}']; c.value=text
    c.font=hf(sz=11); c.fill=fill(color); c.alignment=ctr()
    ws.row_dimensions[row].height=20

def tbl_hdr(ws,row,col,headers):
    for i,h in enumerate(headers):
        c=ws.cell(row=row,column=col+i,value=h)
        c.font=hf(sz=10); c.fill=fill(NAVY); c.alignment=ctr(); c.border=bdr()
    ws.row_dimensions[row].height=18

def data_row(ws,row,col,vals,shade=False):
    for i,v in enumerate(vals):
        c=ws.cell(row=row,column=col+i,value=v)
        c.font=bf(); c.fill=fill(LGRAY if shade else WHITE)
        c.alignment=ctr(); c.border=bdr()
    ws.row_dimensions[row].height=16

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Executive Summary
# ══════════════════════════════════════════════════════════════════════════════
ws1=wb.active; ws1.title='Executive Summary'
ws1.sheet_view.showGridLines=False

for col,w in zip('ABCDEFGH',[2,20,20,20,20,20,20,20]):
    ws1.column_dimensions[col].width=w

# Title
ws1.merge_cells('B1:H2')
ws1['B1']='Patient 30-Day Readmission Risk Analysis'
ws1['B1'].font=Font(name='Arial',bold=True,size=18,color=WHITE)
ws1['B1'].fill=fill(NAVY); ws1['B1'].alignment=ctr()
ws1.row_dimensions[1].height=24; ws1.row_dimensions[2].height=18

ws1.merge_cells('B3:H3')
ws1['B3']='UCI Diabetes 130-US Hospitals Dataset (1999–2008)  |  Strack et al. (2014)  |  CC BY 4.0  |  DOI: 10.24432/C5230J'
ws1['B3'].font=Font(name='Arial',size=9,color=MGRAY,italic=True)
ws1['B3'].fill=fill(NAVY); ws1['B3'].alignment=ctr()
ws1.row_dimensions[3].height=16

ws1.row_dimensions[4].height=10

# KPI cards
kpis=[
    ('Total Patients (unique)',f'{N:,}',NAVY),
    ('30-Day Readmit Rate', pct(overall_rate), RED),
    ('Mean Length of Stay',f"{df['time_in_hospital'].mean():.1f} days",BLUE),
    ('Mean Medications',f"{df['num_medications'].mean():.1f}",DKGRAY),
]
for i,(label,value,color) in enumerate(kpis):
    c1=chr(ord('B')+i*2-1) if i>0 else 'B'
    cols=[('B','C'),('D','E'),('F','G'),('H','H')]
    a,b=cols[i]
    ws1.merge_cells(f'{a}5:{b}6')
    cell=ws1[f'{a}5']
    cell.value=f'{label}\n{value}'
    cell.font=Font(name='Arial',bold=True,size=12,color=WHITE)
    cell.fill=fill(color)
    cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
ws1.row_dimensions[5].height=38; ws1.row_dimensions[6].height=28
ws1.row_dimensions[7].height=10

# Key findings
section_hdr(ws1,8,'B','H','🔍  Key Findings from Real Patient Data — 4 Predictors of 30-Day Readmission')

findings=[
    ('1','Prior Inpatient Visits',
     f'Patients with 5+ prior inpatient visits have a readmit rate of {inpat_stats["mean"].iloc[-1]:.1f}% vs '
     f'{inpat_stats["mean"].iloc[0]:.1f}% for first-time patients — the strongest single predictor.',
     'Patients with frequent prior hospitalizations need intensive post-discharge follow-up'),
    ('2','Diagnosis Category',
     'Circulatory disease patients show the highest volume and elevated readmission rates. '
     'Neoplasms and Diabetes also exceed the overall average.',
     'Chronic disease management programs should prioritise circulatory & diabetes diagnoses'),
    ('3','HbA1c Testing',
     f'Patients with A1C > 8 have the highest readmission rates ({pct(a1c_stats["mean"].loc[">8"])}), '
     'while those with normal results have the lowest — highlighting glycaemic control as a key lever.',
     'Ensure HbA1c testing at every encounter; use results to target discharge interventions'),
    ('4','Length of Stay',
     f'Longer stays correlate with higher readmission: patients staying 12–14 days show '
     f'{pct(los_stats["mean"].iloc[-1])} readmission vs {pct(los_stats["mean"].iloc[0])} for 1–2 day stays.',
     'Extended LOS patients need structured discharge planning and 7-day post-discharge calls'),
]

tbl_hdr(ws1,9,2,['#','Predictor','Finding (Real Data)','Clinical Implication'])
for i,(num,pred,finding,impl) in enumerate(findings):
    shade=i%2==0
    data_row(ws1,10+i,2,[num,pred,finding,impl],shade=shade)

# col widths for table
ws1.column_dimensions['B'].width=4
ws1.column_dimensions['C'].width=22
ws1.column_dimensions['D'].width=55
ws1.column_dimensions['E'].width=45
for r in range(9,14):
    ws1.merge_cells(f'D{r}:E{r}')
    ws1.row_dimensions[r].height=32

ws1.row_dimensions[14].height=10

# Dataset overview table
section_hdr(ws1,15,'B','E','📊  Dataset Overview')
ct_yes=int((df['readmit_30']==1).sum())
ct_gt=int((df['readmitted']=='>30').sum())
ct_no=int((df['readmitted']=='NO').sum())

overview=[
    ('Raw Records (UCI)',          '101,766'),
    ('After Removing Deceased',    '100,114'),
    ('Unique Patients (1/patient)','70,439'),
    ('Readmitted < 30 Days',       f'{ct_yes:,}  ({pct(overall_rate)})'),
    ('Readmitted > 30 Days',       f'{ct_gt:,}  ({ct_gt/N*100:.1f}%)'),
    ('Not Readmitted',             f'{ct_no:,}  ({ct_no/N*100:.1f}%)'),
    ('Most Common Diagnosis',      'Circulatory (30.5%)'),
    ('Mean Length of Stay',        f"{df['time_in_hospital'].mean():.2f} days"),
    ('Mean # Medications',         f"{df['num_medications'].mean():.2f}"),
    ('Mean # Lab Procedures',      f"{df['num_lab_procedures'].mean():.2f}"),
    ('Data Years',                 '1999–2008'),
    ('Hospital Count',             '130 US hospitals'),
    ('Source DOI',                 '10.24432/C5230J'),
]
tbl_hdr(ws1,16,2,['Metric','Value'])
for i,(k,v) in enumerate(overview):
    shade=i%2==0
    ws1.merge_cells(f'B{17+i}:C{17+i}')
    ws1.merge_cells(f'D{17+i}:E{17+i}')
    for col,val in [('B',k),('D',v)]:
        c=ws1[f'{col}{17+i}']
        c.value=val; c.font=bf()
        c.fill=fill(LGRAY if shade else WHITE)
        c.alignment=ctr(); c.border=bdr()
        ws1.row_dimensions[17+i].height=16

ws1.row_dimensions[31].height=10
ws1.merge_cells('B31:H31')
ws1['B31']='Citation: Clore, J., Cios, K., DeShazo, J., & Strack, B. (2014). Diabetes 130-US Hospitals for Years 1999-2008 [Dataset]. UCI Machine Learning Repository. https://doi.org/10.24432/C5230J'
ws1['B31'].font=Font(name='Arial',size=8,color=SLATE,italic=True)
ws1['B31'].alignment=Alignment(horizontal='left',vertical='center')

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Age Analysis
# ══════════════════════════════════════════════════════════════════════════════
ws2=wb.create_sheet('Age Analysis')
ws2.sheet_view.showGridLines=False
for col,w in zip('ABCDEFG',[2,16,16,18,14,14,22]):
    ws2.column_dimensions[col].width=w

ws2.merge_cells('B1:G2')
ws2['B1']='30-Day Readmission Rate by Age Group'
ws2['B1'].font=Font(name='Arial',bold=True,size=14,color=WHITE)
ws2['B1'].fill=fill(NAVY); ws2['B1'].alignment=ctr()
ws2.row_dimensions[1].height=22; ws2.row_dimensions[2].height=16

tbl_hdr(ws2,4,2,['Age Group','Patient Count','% of Total','Readmission Rate','Risk Level','vs. Overall Avg'])
for i,row in age_stats.iterrows():
    rate=row['mean']
    n=row['count']
    pct_total=n/N*100
    diff=rate-overall_rate
    diff_str=f'+{diff:.1f}pp' if diff>=0 else f'{diff:.1f}pp'
    risk='HIGH' if rate>overall_rate*1.3 else ('ELEVATED' if rate>overall_rate else 'LOWER')
    rc=RED if risk=='HIGH' else (AMBER if risk=='ELEVATED' else GREEN)
    shade=list(age_stats.index).index(i)%2==0
    data_row(ws2,5+list(age_stats.index).index(i),2,
             [i,f'{int(n):,}',f'{pct_total:.1f}%',pct(rate),'',diff_str],shade=shade)
    idx=list(age_stats.index).index(i)
    rc_cell=ws2.cell(row=5+idx,column=6)
    rc_cell.value=risk; rc_cell.font=Font(name='Arial',bold=True,size=9,color=WHITE)
    rc_cell.fill=fill(rc); rc_cell.alignment=ctr(); rc_cell.border=bdr()

avg_r=5+len(age_stats)
ws2.merge_cells(f'B{avg_r}:D{avg_r}')
ws2[f'B{avg_r}']='OVERALL AVERAGE (all ages)'
ws2[f'B{avg_r}'].font=Font(name='Arial',bold=True,size=10,color=WHITE)
ws2[f'B{avg_r}'].fill=fill(NAVY); ws2[f'B{avg_r}'].alignment=ctr()
ws2[f'E{avg_r}']=pct(overall_rate)
ws2[f'E{avg_r}'].font=Font(name='Arial',bold=True,size=10,color=WHITE)
ws2[f'E{avg_r}'].fill=fill(NAVY); ws2[f'E{avg_r}'].alignment=ctr()

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — Diagnosis Analysis
# ══════════════════════════════════════════════════════════════════════════════
ws3=wb.create_sheet('Diagnosis Analysis')
ws3.sheet_view.showGridLines=False
for col,w in zip('ABCDEFG',[2,22,16,16,14,14,22]):
    ws3.column_dimensions[col].width=w

ws3.merge_cells('B1:G2')
ws3['B1']='30-Day Readmission Rate by Primary Diagnosis (ICD-9 Mapped)'
ws3['B1'].font=Font(name='Arial',bold=True,size=14,color=WHITE)
ws3['B1'].fill=fill(NAVY); ws3['B1'].alignment=ctr()
ws3.row_dimensions[1].height=22; ws3.row_dimensions[2].height=16

tbl_hdr(ws3,4,2,['Diagnosis Category','Patient Count','% of Total','Readmission Rate','Tier','vs. Overall'])
for i,(cat,row) in enumerate(diag_stats.iterrows()):
    rate=row['mean']; n=row['count']
    pct_t=n/N*100; diff=rate-overall_rate
    diff_str=f'+{diff:.1f}pp' if diff>=0 else f'{diff:.1f}pp'
    tier='Tier 1 – Highest' if i==0 else ('Tier 2' if i<=2 else ('Tier 3' if i<=5 else 'Tier 4 – Lowest'))
    tc=RED if i==0 else (AMBER if i<=2 else (BLUE if i<=5 else GREEN))
    shade=i%2==0
    data_row(ws3,5+i,2,[cat,f'{int(n):,}',f'{pct_t:.1f}%',pct(rate),'',diff_str],shade=shade)
    tc_cell=ws3.cell(row=5+i,column=6)
    tc_cell.value=tier; tc_cell.font=Font(name='Arial',bold=True,size=9,color=WHITE)
    tc_cell.fill=fill(tc); tc_cell.alignment=ctr(); tc_cell.border=bdr()

ws3.row_dimensions[14].height=10
note_row=5+len(diag_stats)+1
ws3.merge_cells(f'B{note_row}:G{note_row}')
ws3[f'B{note_row}']='Note: ICD-9 codes mapped to 9 categories per Strack et al. (2014) Table 2. Primary diagnosis (diag_1) used.'
ws3[f'B{note_row}'].font=Font(name='Arial',size=8,italic=True,color=SLATE)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — Clinical Factors (LOS, Meds, A1C, Prior Visits)
# ══════════════════════════════════════════════════════════════════════════════
ws4=wb.create_sheet('Clinical Factors')
ws4.sheet_view.showGridLines=False
for col,w in zip('ABCDEFGHIJ',[2,18,14,16,20,2,18,14,16,22]):
    ws4.column_dimensions[col].width=w

ws4.merge_cells('B1:E2'); ws4['B1']='Length of Stay'
ws4['B1'].font=Font(name='Arial',bold=True,size=13,color=WHITE)
ws4['B1'].fill=fill(NAVY); ws4['B1'].alignment=ctr()
ws4.row_dimensions[1].height=22; ws4.row_dimensions[2].height=16

tbl_hdr(ws4,4,2,['LOS (days)','Count','Readmit Rate','Interpretation'])
interp=['Typical short stay','Moderate complexity','Elevated risk begins',
        'High complexity','Very high risk','Critical — plan discharge early']
for i,row in los_stats.iterrows():
    shade=list(los_stats.index).index(i)%2==0
    idx=list(los_stats.index).index(i)
    data_row(ws4,5+idx,2,[str(i),f'{int(row["count"]):,}',pct(row['mean']),
                           interp[idx] if idx<len(interp) else ''],shade=shade)

ws4.merge_cells('G1:J2'); ws4['G1']='Medications & Prior Visits'
ws4['G1'].font=Font(name='Arial',bold=True,size=13,color=WHITE)
ws4['G1'].fill=fill(BLUE); ws4['G1'].alignment=ctr()

tbl_hdr(ws4,4,7,['Medications','Count','Readmit Rate','Flag'])
for i,row in med_stats.iterrows():
    idx=list(med_stats.index).index(i)
    flag='⚠ Polypharmacy' if idx>=3 else '—'
    shade=idx%2==0
    data_row(ws4,5+idx,7,[str(i),f'{int(row["count"]):,}',pct(row['mean']),flag],shade=shade)
    if idx>=3:
        ws4.cell(row=5+idx,column=10).font=Font(name='Arial',color=RED,bold=True,size=10)

# A1C sub-table
ws4.row_dimensions[12].height=10
section_hdr(ws4,13,'B','E','HbA1c Result vs 30-Day Readmission',AMBER)
tbl_hdr(ws4,14,2,['A1C Result','Count','Readmit Rate','Clinical Note'])
a1c_notes=['A1C not measured — incomplete glycaemic data',
           'Controlled glycaemia — lowest readmit risk',
           'Above target — moderate risk elevation',
           'Poorly controlled — highest readmit risk']
for i,row in a1c_stats.iterrows():
    idx=list(a1c_stats.index).index(i)
    shade=idx%2==0
    n_val=int(row['count']) if not np.isnan(row['count']) else 0
    data_row(ws4,15+idx,2,[str(i),f'{n_val:,}',pct(row['mean']),
                            a1c_notes[idx] if idx<len(a1c_notes) else ''],shade=shade)

# Prior inpatient visits sub-table
section_hdr(ws4,20,'G','J','Prior Inpatient Visits vs Readmission',NAVY)
tbl_hdr(ws4,21,7,['Prior Visits','Count','Readmit Rate','Risk Impact'])
for i,row in inpat_stats.iterrows():
    idx=list(inpat_stats.index).index(i)
    lbl=str(int(i)) if i<5 else '5+'
    shade=idx%2==0
    impact=f'+{row["mean"]-overall_rate:.1f}pp vs avg' if row['mean']>overall_rate else f'{row["mean"]-overall_rate:.1f}pp vs avg'
    data_row(ws4,22+idx,7,[lbl,f'{int(row["count"]):,}',pct(row['mean']),impact],shade=shade)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — Visualizations
# ══════════════════════════════════════════════════════════════════════════════
ws5=wb.create_sheet('Visualizations')
ws5.sheet_view.showGridLines=False
ws5.merge_cells('A1:N2'); ws5['A1']='Matplotlib Visualizations — Real UCI Dataset'
ws5['A1'].font=Font(name='Arial',bold=True,size=15,color=WHITE)
ws5['A1'].fill=fill(NAVY); ws5['A1'].alignment=ctr()
ws5.row_dimensions[1].height=22; ws5.row_dimensions[2].height=16

for fig_path,anchor,label,label_row,h in [
    ('fig_key_predictors.png','A4','Figure 1 — Four Key Predictors (Age, Diagnosis, LOS, Prior Visits)',3,590),
    ('fig_dashboard.png','A35','Figure 2 — Full Analysis Dashboard (Race, HbA1c, LOS Distribution, Bubble Chart)',34,520),
]:
    ws5.merge_cells(f'A{label_row}:N{label_row}')
    lc=ws5[f'A{label_row}']; lc.value=label
    lc.font=Font(name='Arial',bold=True,size=11,color=WHITE)
    lc.fill=fill(BLUE); lc.alignment=ctr(); ws5.row_dimensions[label_row].height=18
    img=XLImage(fig_path); img.width=920; img.height=h
    ws5.add_image(img,anchor)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 6 — Data Sample
# ══════════════════════════════════════════════════════════════════════════════
ws6=wb.create_sheet('Data Sample (500 rows)')
ws6.sheet_view.showGridLines=False

keep_cols=['encounter_id','patient_nbr','race','gender','age','time_in_hospital',
           'num_lab_procedures','num_procedures','num_medications','number_outpatient',
           'number_emergency','number_inpatient','number_diagnoses',
           'A1Cresult','insulin','change','diabetesMed','diag_cat','readmitted','readmit_30']
sample=df[keep_cols].head(500)

for j,col in enumerate(sample.columns):
    c=ws6.cell(row=1,column=j+1,value=col.replace('_',' ').title())
    c.font=hf(sz=10); c.fill=fill(NAVY); c.alignment=ctr(); c.border=bdr()
    ws6.column_dimensions[get_column_letter(j+1)].width=18
ws6.row_dimensions[1].height=18

for i,(_,row) in enumerate(sample.iterrows()):
    shade=i%2==0
    for j,val in enumerate(row):
        c=ws6.cell(row=i+2,column=j+1,value=val)
        c.font=bf(sz=9); c.fill=fill(LGRAY if shade else WHITE)
        c.alignment=ctr(); c.border=bdr('hair')
    ws6.row_dimensions[i+2].height=14
ws6.freeze_panes='A2'

# Tab colours
ws1.sheet_properties.tabColor=NAVY
ws2.sheet_properties.tabColor=BLUE
ws3.sheet_properties.tabColor=RED
ws4.sheet_properties.tabColor=AMBER
ws5.sheet_properties.tabColor='7C3AED'
ws6.sheet_properties.tabColor=SLATE

out='Patient_Readmission_Risk_Analysis_REAL.xlsx'
wb.save(out)
print(f"Saved: {out}")
